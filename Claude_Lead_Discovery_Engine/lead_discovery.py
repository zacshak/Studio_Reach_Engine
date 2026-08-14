"""SYSTEM 1 - LEAD DISCOVERY  (Finder + Fetcher in one command).

Produces a clean, browsable list of NEW pre-release Steam studios:
  game name, studio, genre, brief, release status, Steam page, website,
  and Steam's own support email.  (Email/Country/Engine enrichment = System 2.)

Output: out/leads_<date>.xlsx  AND  out/leads_<date>.csv

Three ways to run it:
  python lead_discovery.py --sample 25     # SEE DATA NOW: 25 random pre-release studios
  python lead_discovery.py                 # daily: process today's newly-found pages
  python lead_discovery.py --file out/new_appids_2026-06-20.txt   # a specific list

(The daily mode reads out/new_appids_<today>.txt, which `find_new.py` produces.)
"""
import argparse
import csv
import glob
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import batch_fetch as bf  # reuse the cached, rate-limited fetcher
import pipeline  # shared DB connection (Turso when configured)
from approval_export import ApprovalExporter, NOMAIL_BASE  # background staging

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# (header, key in normalized dict, column width)
COLUMNS = [
    ("AppID", "appid", 10),
    ("Game Name", "game_name", 30),
    ("Type", "type", 8),
    ("Status", "_status", 16),
    ("Studio Name", "_developer", 26),
    ("Publisher", "_publisher", 22),
    ("Genre", "_genres", 22),
    ("Steam Page", "page_link", 42),
    ("Studio Website", "studio_website", 38),
    ("Steam Support Email", "support_email", 28),
    ("Brief", "brief", 70),
]


def to_row(n):
    r = dict(n)
    r["_developer"] = ", ".join(n.get("developers") or []) or "—"
    r["_publisher"] = ", ".join(n.get("publishers") or []) or "—"
    r["_genres"] = ", ".join(n.get("genres") or []) or "—"
    rd = n.get("release_date") or ""
    r["_status"] = "Coming Soon" if n.get("coming_soon") else (rd or "released?")
    return {h: ("" if r.get(k) is None else r.get(k, "")) for h, k, _ in COLUMNS}


def load_appids(args, conn, stamp):
    if args.sample:
        rows = conn.execute(
            "SELECT appid FROM known_comingsoon ORDER BY RANDOM() LIMIT ?",
            (args.sample,)).fetchall()
        return [r[0] for r in rows], f"random sample of {args.sample} pre-release pages"
    if args.file:
        path = args.file
    elif args.stamp:
        # runner mode: the finder wrote new_appids_<stamp>.txt this same run
        path = os.path.join(bf.OUT_DIR, f"new_appids_{stamp}.txt")
    else:
        # standalone: use the most recent finder output
        candidates = sorted(glob.glob(os.path.join(bf.OUT_DIR, "new_appids_*.txt")))
        if not candidates:
            return None, ("no finder output found in out/.\n"
                          "   Run the finder first:  python find_new.py\n"
                          "   Or see data now:        python lead_discovery.py --sample 25")
        path = candidates[-1]
    if not os.path.exists(path):
        return None, (f"no input list found at {path}\n"
                      f"   Run the finder first:  python find_new.py\n"
                      f"   Or see data now:        python lead_discovery.py --sample 25")
    ids = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                ids.append(int(line.split("#")[0].split()[0]))
    return ids, f"newly-found pages ({os.path.basename(path)})"


def write_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    for ci, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, header)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width
    for ri, row in enumerate(rows, 2):
        for ci, (header, _, _) in enumerate(COLUMNS, 1):
            val = row[header]
            cell = ws.cell(ri, ci, val)
            if header == "Brief":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sample", type=int, help="see data now: N random pre-release studios")
    ap.add_argument("--file", help="a specific new_appids file")
    ap.add_argument("--stamp", help="shared run stamp (YYYY-MM-DD_HHMM) from the runner")
    ap.add_argument("--refresh", action="store_true", help="ignore cache, re-fetch")
    ap.add_argument("--include-released", action="store_true",
                    help="keep released games too (default: pre-release only)")
    args = ap.parse_args()

    os.makedirs(bf.OUT_DIR, exist_ok=True)
    conn = pipeline.connect()
    bf.init_db(conn)
    pipeline.init_tracker()

    today = datetime.now(bf.IST).date().isoformat()
    stamp = args.stamp or datetime.now(bf.IST).strftime("%Y-%m-%d_%H%M")

    appids, source = load_appids(args, conn, stamp)
    if appids is None:
        print(f"\n  Nothing to do: {source}\n")
        return 1
    print(f"Lead Discovery on {len(appids)} appids — {source}\n")

    # --sample is a PREVIEW tool: it must never write into newly_added.
    # Only real runs (a diff list or an explicit --file) persist leads.
    is_sample = bool(args.sample)
    # Exporters stage folders + screenshots concurrently after source validation.
    exporter = None if is_sample else ApprovalExporter()
    # same staging for no-email leads, into No_Mail_Games (scrape_status 'pending')
    nomail_exporter = None if is_sample else ApprovalExporter(base_dir=NOMAIL_BASE,
                                                              status="pending")
    limiter = bf.RateLimiter(bf.MIN_INTERVAL, bf.MAX_PER_WINDOW, bf.WINDOW)
    rows, tracked_appids, ok, dropped = [], [], 0, 0
    for idx, appid in enumerate(appids, 1):
        hit = None if args.refresh else bf.cache_get(conn, appid)
        if hit is not None:
            norm = hit["normalized"] if hit["success"] else None
            tracked = norm is not None
            tag = "cache"
        else:
            success, norm, raw = bf.fetch_one(appid, limiter)
            tracked = False
            if not is_sample:  # cache_put itself guards to pre-release leads only
                tracked = bool(bf.cache_put(
                    conn, appid, success, raw, discovered_on=today) and norm)
            tag = "fetched"
        if not norm:
            dropped += 1
            continue
        # keep only games + demos; pre-release only unless overridden
        if norm.get("type") not in ("game", "demo"):
            dropped += 1
            continue
        if not args.include_released and not norm.get("coming_soon"):
            dropped += 1
            continue
        if not is_sample and tracked:
            tracked_appids.append(appid)
        rows.append(to_row(norm))
        ok += 1
        print(f"[{idx}/{len(appids)}] {tag:<7} {norm['game_name'][:42]}  "
              f"[{norm['type']}, {'coming soon' if norm['coming_soon'] else 'released'}]")

    if not is_sample:
        # Reconcile with the shared validator before any media exporter can stage a
        # lead. This is the first point at which source contact data becomes a queue
        # state, so invalid values never reach the outreach review folder.
        pipeline.seed_pending()
        invalid = 0
        for appid in tracked_appids:
            norm = pipeline.read_lead(appid) or {}
            status = norm.get("support_info")
            state = pipeline.discovery_status(pipeline._support_email(status))
            if state == "seeded":
                exporter.submit(appid)
            elif state == "invalid":
                invalid += 1
            else:
                nomail_exporter.submit(appid)
    else:
        invalid = 0

    csv_path = os.path.join(bf.OUT_DIR, f"leads_{stamp}.csv")
    xlsx_path = os.path.join(bf.OUT_DIR, f"leads_{stamp}.xlsx")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[h for h, _, _ in COLUMNS])
        w.writeheader()
        w.writerows(rows)
    write_excel(rows, xlsx_path)

    # wait for concurrent exports to finish
    staged, exp_errors = exporter.close() if exporter else (0, [])
    nomail_staged, nomail_errors = nomail_exporter.close() if nomail_exporter else (0, [])

    print("\n" + "=" * 60)
    print("  LEAD DISCOVERY COMPLETE")
    print(f"  {ok} pre-release leads kept, {dropped} skipped (released/no-data/other)")
    print(f"  Excel: {xlsx_path}")
    print(f"  CSV  : {csv_path}")
    if exporter:
        print(f"  Approval-staged: {staged} seeded lead(s)"
              + (f", {len(exp_errors)} export error(s)" if exp_errors else ""))
        print(f"  Invalid         : {invalid} lead(s) quarantined")
        print(f"  No-mail-staged : {nomail_staged} pending lead(s)"
              + (f", {len(nomail_errors)} export error(s)" if nomail_errors else ""))
    print("=" * 60)
    if rows:
        print("\n  Preview (first 8):")
        print(f"  {'Game':<32}{'Studio':<24}{'Status':<14}")
        print("  " + "-" * 68)
        for r in rows[:8]:
            print(f"  {r['Game Name'][:30]:<32}{r['Studio Name'][:22]:<24}{r['Status'][:12]:<14}")
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
